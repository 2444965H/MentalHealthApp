#This file loads the input and then parses it into a class for the internal representation (IR)
#The classes are divided into subsections, each with its own IR array.
#Print statements (that are hashed out) can be unhashed in order to check for correct array (output)
#The classes' IR arrays are then processed in the next step in the IRProcessor.py

from openpyxl import load_workbook

#import openpyxl
#workbook = openpyxl.open_workbook('file.xlsx')
#worksheet = workbook.sheet_by_index(0)
#print(worksheet.cell(0, 0).value)
wb = load_workbook(filename ='Input2.xlsm')
ws = wb['Day 1']
#Tab namens "test" - Klein/Großschreibung achten!

class DepressionIR:
    DepressionArray = []
    for x in range(4,13):
        DepressionArray.append(ws.cell(row=x, column=2).value)
        #print(ws.cell(row=x, column=2).value)
    #print(len(DepressionArray))

class AnxietyIR:
    AnxietyArray = []
    for x in range(18,25):
        AnxietyArray.append(ws.cell(row=x, column=2).value)
        #print(ws.cell(row=x, column=2).value)
    #print(len(AnxietyArray))

class StressIR:
    StressArray = []
    for x in range(29,39):
        StressArray.append(ws.cell(row=x, column=2).value)
        #print(ws.cell(row=x, column=2).value)
    #print(len(StressArray))

